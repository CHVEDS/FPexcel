"""
Парсер Excel с вычислением формул, извлечением изображений и метаданных ячеек
Использует:
xlwings: вычисление формул через Excel (поддержка русских функций, внешних ссылок)
openpyxl: извлечение изображений, формул, форматирования, объединённых ячеек
Автор: Чураев Вадим Эдуардович || РЭУ ИМ.ПЛЕХАНОВА
"""
import os
import re
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl_image_loader import SheetImageLoader

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(name)

# ---------------------------------------------------------------------------
# Вспомогательные функции
# ---------------------------------------------------------------------------
def _col_letter(n: int) -> str:
    return get_column_letter(n)

def _resolve_merged(ws_op) -> Dict[str, str]:
    """Возвращает сопоставление координат -> координаты левого верхнего угла для каждой ячейки в объединённом диапазоне."""
    mapping: Dict[str, str] = {}
    for merged_range in ws_op.merged_cells.ranges:
        top_left = f"{_col_letter(merged_range.min_col)}{merged_range.min_row}"
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                coord = f"{_col_letter(col)}{row}"
                if coord != top_left:
                    mapping[coord] = top_left
    return mapping

def _detect_real_bounds(ws_op, max_scan_col: int = 500) -> Tuple[int, int]:
    """
    Определяет фактическую последнюю строку/столбец с данными, ограничиваясь max_scan_col столбцами.
    Избегает раздутого used_range, который openpyxl отчитывает при наличии случайных стилей
    далеко справа (например, A1:XCA20323, когда данные заканчиваются на BN).
    """
    last_row = 0
    last_col = 0
    for row in ws_op.iter_rows(max_col=max_scan_col):
        for cell in row:
            if cell.value is not None:
                if cell.row > last_row:
                    last_row = cell.row
                if cell.column > last_col:
                    last_col = cell.column
    return last_row, last_col

def _cell_color(cell) -> Optional[str]:
    """Возвращает hex-код цвета заливки ячейки или None, если заливка отсутствует/прозрачна."""
    try:
        fg = cell.fill.fgColor
        if fg.type == "rgb" and fg.rgb not in ("00000000", "FFFFFFFF", "00FFFFFF"):
            return fg.rgb
        if fg.type == "theme":
            return f"theme:{fg.theme}"
    except Exception:
        pass
    return None

def _cell_formatting(cell) -> Dict[str, Any]:
    fmt: Dict[str, Any] = {}
    try:
        fmt["bold"] = bool(cell.font.bold)
        fmt["font_name"] = cell.font.name
        fmt["font_size"] = cell.font.size
    except Exception:
        pass
    color = _cell_color(cell)
    if color:
        fmt["fill_color"] = color
    try:
        fmt["number_format"] = cell.number_format
    except Exception:
        pass
    return fmt

# ---------------------------------------------------------------------------
# Основной парсер
# ---------------------------------------------------------------------------
def parse_excel(
    filepath: str,
    image_dir: str = "images",
    sheet_name: Optional[str] = None,
    close_app: bool = True,
    skip_empty: bool = True,
    include_formula: bool = True,
    include_formatting: bool = False,
    max_col: Optional[int] = None,
    max_row: Optional[int] = None,
) -> Dict[str, Dict[str, Any]]:
    """
    Парсит Excel-файл и извлекает вычисленные значения, формулы, изображения,
    информацию об объединённых ячейках и, при необходимости, форматирование.

    Параметры
    ----------
    filepath : str
        Путь к файлу .xlsx.
    image_dir : str
        Директория для сохранения извлечённых изображений.
    sheet_name : str, optional
        Имя листа или None для активного листа.
    close_app : bool
        Закрыть Excel после чтения (по умолчанию True).
    skip_empty : bool
        Пропускать ячейки, где значение, формула и изображение равны None/пусты.
    include_formula : bool
        Включить исходный текст формулы в результат (по умолчанию True).
    include_formatting : bool
        Включить метаданные форматирования ячейки (жирный, цвет заливки и т.д.).
    max_col : int, optional
        Остановить сканирование после этого номера столбца. Полезно для файлов с
        тысячами лишних столбцов (например, случайные стили до XCA).
        Если None, автоматически определяет границы, сканируя до 500 столбца.
    max_row : int, optional
        Остановить сканирование после этой строки. Если None, использует все строки.

    Возвращает
    -------
    dict
        Ключи — координаты Excel (например, "A1"):
        - "value"      : вычисленное значение ячейки (str/int/float/datetime/None)
        - "formula"    : строка формулы или None (если include_formula)
        - "image_path" : путь к сохранённому изображению или None
        - "merged_into": координата левой верхней ячейки объединения, если применимо
        - "formatting" : словарь с bold/fill_color/etc. (если include_formatting)

    Исключения
    ------
    FileNotFoundError
        Если Excel-файл не найден.
    RuntimeError
        Если не удалось открыть файл в Excel.
    """
    filepath = os.path.abspath(filepath)
    if not os.path.isfile(filepath):
        raise FileNotFoundError(f"Excel-файл не найден: {filepath}")

    Path(image_dir).mkdir(parents=True, exist_ok=True)
    logger.info(f"Парсинг: {filepath}")

    # ── Шаг 1: вычисление значений через xlwings ──────────────────────────────────
    computed_values: Dict[str, Any] = {}
    app = None
    wb_xw = None
    try:
        logger.debug("Открытие Excel через xlwings...")
        app = xw.App(visible=False, add_book=False)
        wb_xw = app.books.open(filepath)
        ws_xw = wb_xw.sheets[sheet_name] if sheet_name else wb_xw.sheets.active

        used = ws_xw.used_range
        if used is None:
            logger.warning("xlwings: используемый диапазон не обнаружен.")
        else:
            raw = used.value
            if raw is None:
                logger.warning("xlwings: используемый диапазон пуст.")
            else:
                if not isinstance(raw, list):
                    raw = [[raw]]
                elif not isinstance(raw[0], list):
                    raw = [raw]

                origin_row = used.row
                origin_col = used.column
                for i, row in enumerate(raw):
                    for j, val in enumerate(row):
                        r = origin_row + i
                        c = origin_col + j
                        # Учитываем ограничения max_col / max_row
                        if max_col and c > max_col:
                            break
                        if max_row and r > max_row:
                            break
                        coord = f"{_col_letter(c)}{r}"
                        computed_values[coord] = val

    except Exception as e:
        logger.exception("Ошибка xlwings — вычисленные значения недоступны")
        raise RuntimeError(f"Ошибка xlwings: {e}") from e
    finally:
        if wb_xw:
            try:
                wb_xw.close()
            except Exception:
                pass
        if app and close_app:
            try:
                app.quit()
            except Exception:
                pass

    # ── Шаг 2: формулы, изображения, форматирование через openpyxl ───────────────────
    logger.debug("Загрузка книги через openpyxl...")
    wb_op = load_workbook(filepath, data_only=False)
    ws_op = wb_op[sheet_name] if sheet_name else wb_op.active

    # Автоопределение реальных границ столбцов/строк для пропуска "раздутых" областей
    if max_col is None:
        logger.debug("Автоопределение реальных границ столбцов (макс. 500)...")
        _, detected_col = _detect_real_bounds(ws_op, max_scan_col=500)
        effective_max_col = max(detected_col, 1) if detected_col else 500
    else:
        effective_max_col = max_col

    effective_max_row = max_row  # None = без ограничений

    # Сопоставление объединённых ячеек
    merged_map = _resolve_merged(ws_op)

    # Загрузчик изображений
    try:
        image_loader = SheetImageLoader(ws_op)
    except (ValueError, Exception) as e:
        logger.warning(f"Загрузчик изображений недоступен: {e}")
        image_loader = None

    result: Dict[str, Dict[str, Any]] = {}

    for row in ws_op.iter_rows(max_col=effective_max_col, max_row=effective_max_row):
        for cell in row:
            coord = cell.coordinate

            # Формула
            formula: Optional[str] = None
            if include_formula and isinstance(cell.value, str) and cell.value.startswith("="):
                formula = cell.value

            # Изображение
            img_path: Optional[str] = None
            if image_loader:
                try:
                    if image_loader.image_in(coord):
                        img = image_loader.get(coord)
                        img_path = os.path.join(image_dir, f"{coord}.png")
                        img.save(img_path)
                        logger.debug(f"Изображение сохранено: {coord} -> {img_path}")
                except Exception as img_err:
                    logger.error(f"Ошибка изображения в {coord}: {img_err}")

            value = computed_values.get(coord)

            # Пропуск полностью пустых ячеек
            if skip_empty and value is None and formula is None and img_path is None:
                continue

            entry: Dict[str, Any] = {
                "value": value,
                "image_path": img_path,
            }

            if include_formula:
                entry["formula"] = formula

            merged_into = merged_map.get(coord)
            if merged_into:
                entry["merged_into"] = merged_into

            if include_formatting:
                entry["formatting"] = _cell_formatting(cell)

            result[coord] = entry

    wb_op.close()
    logger.info(f"Готово. Обработано {len(result)} ячеек.")
    return result

# ---------------------------------------------------------------------------
# Вспомогательная функция для извлечения заголовков
# ---------------------------------------------------------------------------
def extract_headers(
    result: Dict[str, Dict[str, Any]],
    header_rows: List[int],
) -> Dict[str, str]:
    """
    Создаёт сопоставление столбец -> заголовок на основе одной или нескольких строк заголовков.
    Конкатенирует непустые значения из каждой строки заголовка для каждого столбца.
    Пример: header_rows=[2, 3] имитирует паттерн из двух строк (группа/подзаголовок)
    как в Практика Python.xlsx.
    """
    headers: Dict[str, List[str]] = {}
    for coord, info in result.items():
        match = re.match(r"([A-Z]+)(\d+)", coord)
        if not match:
            continue
        col, row = match.group(1), int(match.group(2))
        if row not in header_rows:
            continue
        val = info.get("value")
        if val is None:
            val = info.get("formula") or ""
        val = str(val).strip()
        if val:
            headers.setdefault(col, []).append(val)

    return {col: " / ".join(parts) for col, parts in headers.items()}

# ---------------------------------------------------------------------------
# Точка входа CLI
# ---------------------------------------------------------------------------
if name == "main":
    import sys
    if len(sys.argv) < 2:
        print("Использование: python \"parser (1).py\" <файл.xlsx> [директория_изображений] [имя_листа]")
        sys.exit(1)

    excel_path = sys.argv[1]
    img_dir    = sys.argv[2] if len(sys.argv) > 2 else "images"
    sname      = sys.argv[3] if len(sys.argv) > 3 else None

    try:
        data = parse_excel(
            excel_path,
            image_dir=img_dir,
            sheet_name=sname,
            include_formula=True,
            include_formatting=True,
        )

        # Вывод сводки
        has_image = sum(1 for v in data.values() if v.get("image_path"))
        has_formula = sum(1 for v in data.values() if v.get("formula"))
        print(f"Всего ячеек: {len(data)}")
        print(f"  со значениями : {sum(1 for v in data.values() if v['value'] is not None)}")
        print(f"  с формулами: {has_formula}")
        print(f"  с изображениями  : {has_image}")

        # Вывод первых 30 непустых ячеек
        print("\nПервые 30 ячеек:")
        for coord, info in list(sorted(data.items()))[:30]:
            print(f"  {coord}: значение={repr(info['value'])[:50]}  формула={info.get('formula','')[:40]}")

    except Exception:
        logger.exception("Ошибка парсинга")
        sys.exit(1)
